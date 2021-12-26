import openpyxl as xl
from openpyxl import chart
from openpyxl.chart import BarChart, Reference

# def process_workbook(filename):
wb = xl.load_workbook('transaction.xlsx');
sheet = wb['Sheet1']

for rows in range(2, sheet.max_row +1):
    cell = sheet.cell(rows, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(rows, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'a13')
wb.save('filename2.xlsx')
    
    
    
    
